import glob
import numpy as np
import pandas as pd
from pandas.core.frame import DataFrame
from sklearn.preprocessing import StandardScaler
from sklearn.cross_decomposition import PLSRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
import openpyxl

def polyfit(x, y, degree):
    results = {}
    coeffs = np.polyfit(x, y, degree)
    results['polynomial'] = coeffs.tolist()
    p = np.poly1d(coeffs)
    yhat = p(x)
    ybar = np.sum(y) / len(y)
    ssreg = np.sum((yhat - ybar) ** 2)
    sstot = np.sum((y - ybar) ** 2)
    results['determination'] = ssreg / sstot
    return results

land_cover=['0_NDVI_0_1','0_NDVI_0_2','0_NDVI_0_3','0_NDVI_0_4','0_NDVI_0_5','0_NDVI_0_6','0_NDVI_0_7','0_NDVI_0_8','0_NDVI_0_9']


book=openpyxl.Workbook()
sheet=book.create_sheet(index=0,title='result of RF')
sheet.cell(column=1, row=1, value='R')
sheet.cell(column=2, row=1, value='R1')
#sheet.cell(column=2, row=1, value='VIP_PAR')
#sheet.cell(column=3, row=1, value='VIP_TMP')
#sheet.cell(column=4, row=1, value='VIP_PRE')
#sheet.cell(column=5, row=1, value='VIP_VPD')
#sheet.cell(column=6, row=1, value='VIP_SM')
'''
sheet1=book.create_sheet(index=1,title='NDVI_1')
sheet1.cell(column=1, row=1, value='Test_DATA')
sheet1.cell(column=2, row=1, value='Predict')

sheet2=book.create_sheet(index=2,title='NDVI_2')
sheet2.cell(column=1, row=1, value='Test_DATA')
sheet2.cell(column=2, row=1, value='Predict')

sheet3=book.create_sheet(index=3,title='NDVI_3')
sheet3.cell(column=1, row=1, value='Test_DATA')
sheet3.cell(column=2, row=1, value='Predict')

sheet4=book.create_sheet(index=4,title='NDVI_4')
sheet4.cell(column=1, row=1, value='Test_DATA')
sheet4.cell(column=2, row=1, value='Predict')

sheet5=book.create_sheet(index=5,title='NDVI_5')
sheet5.cell(column=1, row=1, value='Test_DATA')
sheet5.cell(column=2, row=1, value='Predict')

sheet6=book.create_sheet(index=6,title='NDVI_6')
sheet6.cell(column=1, row=1, value='Test_DATA')
sheet6.cell(column=2, row=1, value='Predict')

sheet7=book.create_sheet(index=7,title='NDVI_7')
sheet7.cell(column=1, row=1, value='Test_DATA')
sheet7.cell(column=2, row=1, value='Predict')

sheet8=book.create_sheet(index=8,title='NDVI_8')
sheet8.cell(column=1, row=1, value='Test_DATA')
sheet8.cell(column=2, row=1, value='Predict')

sheet9=book.create_sheet(index=9,title='NDVI_9')
sheet9.cell(column=1, row=1, value='Test_DATA')
sheet9.cell(column=2, row=1, value='Predict')
'''
for land_index in range(0,9):
    data = pd.read_csv('E:/1_canopy_photosynthesis/photosynthesis_and_Global_Climate/SIF/data/meoto/TEST/Points/'+land_cover[land_index]+'/FLUXCOM_CSIF_1.csv',header=0, dtype=np.float64)
    data1 = pd.read_csv('E:/1_canopy_photosynthesis/photosynthesis_and_Global_Climate/SIF/data/meoto/TEST/Points/'+land_cover[land_index]+'/FLUXCOM_CSIF_2.csv',header=0, dtype=np.float64)

    data = np.array(data)

    data1 = np.array(data1)
    m, n = data1.shape
    pointID_1 = data[:, 1].tolist()
    pointID_2 = data1[:, 1].tolist()
    data_t = np.zeros((m, n))
    for i in range(len(pointID_1)):
        index = pointID_1.index(pointID_2[i])
        data_t[index, :] = data1[i, :]
    data1 = data_t

    kk = 192
    FID = data[:, 0]
    data_1 = np.where(data[:, 3:3 + kk] > 0, data[:, 3:3 + kk], None)
    data_2 = np.where(data[:, 3 + kk:3 + 2 * kk] > 0, data[:, 3 + kk:3 + 2 * kk], None)
    data_3 = np.where(data[:, 3 + 2 * kk:3 + 3 * kk] > 0, data[:, 3 + 2 * kk:3 + 3 * kk], None)
    data_4 = np.where(data[:, 3 + 3 * kk:3 + 4 * kk] >= -100, data[:, 3 + 3 * kk:3 + 4 * kk], None)

    #data_5 = np.where(data[:, 3 + 4 * kk:3 + 5 * kk] >= 0, data[:, 3 + 4 * kk:3 + 5 * kk], None)
    #data_6 = np.where(data[:, 3 + 5 * kk:3 + 6 * kk] >= 0, data[:, 3 + 5 * kk:3 + 6 * kk], None)

    data_5 = np.where(data1[:, 3 + kk:3 + 2 * kk] >= 0, data1[:, 3 + kk:3 + 2 * kk], None)
    data_6 = np.where(data1[:, 3 + 2 * kk:3 + 3 * kk] >= 0, data1[:, 3 + 2 * kk:3 + 3 * kk], None)
    data_7 = np.where(data1[:, 3 + 3 * kk:3 + 4 * kk] >= 0, data1[:, 3 + 3 * kk:3 + 4 * kk], None)

    listGPP = data_1.ravel().tolist()
    listSIF = data_2.ravel().tolist()
    listPAR = data_3.ravel().tolist()
    listTMP = data_4.ravel().tolist()
    listPRE = data_5.ravel().tolist()
    listVPD = data_6.ravel().tolist()
    listSM = data_7.ravel().tolist()

    list_GPP=[]
    list_SIF=[]
    list_PAR=[]
    list_TMP=[]
    list_PRE=[]
    list_VPD=[]
    list_SM=[]

    for j in range(0,len(listGPP)):
        if listGPP[j] != None and listSIF[j] != None and listPAR[j] != None and listTMP[j] != None and listPRE[j] != None and listVPD[j]!=None and listSM[j]!=None:
            list_GPP.append(listGPP[j])
            list_SIF.append(listSIF[j])
            list_PAR.append(listPAR[j])
            list_TMP.append(listTMP[j])
            list_PRE.append(listPRE[j])
            list_VPD.append(listVPD[j])
            list_SM.append(listSM[j])

    x = {"list_PAR": list_PAR, "list_TMP": list_TMP, "list_PRE": list_PRE,"list_VPD":list_VPD,"list_SM": list_SM}
    y = {"list_GPP": list_GPP,'list_SIF':list_SIF}

    DATAx = DataFrame(x)
    DATAy = DataFrame(y)
    X_train, X_test, y_train, y_test = train_test_split(DATAx,DATAy,test_size=0.2,random_state=0)

    sc=StandardScaler()
    X_train=sc.fit_transform(X_train)
    X_test=sc.fit_transform(X_test)

    regr=RandomForestRegressor(n_estimators=50,random_state=0)
    regr.fit(X_train,y_train)
    #score=regr.score(X_test,y_test)
    y_pred=regr.predict(X_test)
    R = polyfit(y_test['list_GPP'], y_pred[:,0], 1)['determination']
    R1 = polyfit(y_test['list_SIF'], y_pred[:,1], 1)['determination']
    importances=regr.feature_importances_
    '''
    if land_index==0:
        for i in range(0,len(y_pred)):
            sheet1.cell(column=1, row=i + 2, value=np.array(y_test).ravel()[i])
            sheet1.cell(column=2, row=i + 2, value=y_pred[i])
    elif land_index==1:
        for i in range(0,len(y_pred)):
            sheet2.cell(column=1, row=i + 2, value=np.array(y_test).ravel()[i])
            sheet2.cell(column=2, row=i + 2, value=y_pred[i])
    elif land_index==2:
        for i in range(0,len(y_pred)):
            sheet3.cell(column=1, row=i + 2, value=np.array(y_test).ravel()[i])
            sheet3.cell(column=2, row=i + 2, value=y_pred[i])
    elif land_index==3:
        for i in range(0,len(y_pred)):
            sheet4.cell(column=1, row=i + 2, value=np.array(y_test).ravel()[i])
            sheet4.cell(column=2, row=i + 2, value=y_pred[i])
    elif land_index==4:
        for i in range(0,len(y_pred)):
            sheet5.cell(column=1, row=i + 2, value=np.array(y_test).ravel()[i])
            sheet5.cell(column=2, row=i + 2, value=y_pred[i])
    elif land_index==5:
        for i in range(0,len(y_pred)):
            sheet6.cell(column=1, row=i + 2, value=np.array(y_test).ravel()[i])
            sheet6.cell(column=2, row=i + 2, value=y_pred[i])
    elif land_index==6:
        for i in range(0,len(y_pred)):
            sheet7.cell(column=1, row=i + 2, value=np.array(y_test).ravel()[i])
            sheet7.cell(column=2, row=i + 2, value=y_pred[i])
    elif land_index==7:
        for i in range(0,len(y_pred)):
            sheet8.cell(column=1, row=i + 2, value=np.array(y_test).ravel()[i])
            sheet8.cell(column=2, row=i + 2, value=y_pred[i])
    elif land_index==8:
        for i in range(0,len(y_pred)):
            sheet9.cell(column=1, row=i + 2, value=np.array(y_test).ravel()[i])
            sheet9.cell(column=2, row=i + 2, value=y_pred[i])
    '''
    sheet.cell(column=1, row=land_index + 2, value=R)
    sheet.cell(column=2, row=land_index + 2, value=R1)
    #sheet.cell(column=2, row=land_index + 2, value=importances[0])
    #sheet.cell(column=3, row=land_index + 2, value=importances[1])
    #sheet.cell(column=4, row=land_index + 2, value=importances[2])
    #sheet.cell(column=5, row=land_index + 2, value=importances[3])
    #sheet.cell(column=6, row=land_index + 2, value=importances[4])
book.save(r'C:\Users\lijia\Desktop\FLUXCOM_CSIF.xlsx')

print()
