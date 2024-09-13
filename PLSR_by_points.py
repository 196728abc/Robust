
import glob
import numpy as np
import pandas as pd
import math
from pandas.core.frame import DataFrame
import pingouin as pg
from sklearn.linear_model import LinearRegression
import openpyxl as op
from sklearn import preprocessing
import openpyxl
from sklearn.preprocessing import StandardScaler
from sklearn.cross_decomposition import PLSRegression
from sklearn.model_selection import train_test_split


def polyfit(x,y,degree):
    results = {}
    coeffs = np.polyfit(x,y,degree)
    results['polynomial'] = coeffs.tolist()
    p = np.poly1d(coeffs)
    yhat = p(x)
    ybar = np.sum(y)/len(y)
    ssreg = np.sum((yhat - ybar)**2)
    sstot = np.sum((y - ybar)**2)
    results['determination'] = ssreg/sstot

    rmse=math.sqrt(np.sum((y-yhat)**2)/len(y))
    results['RMSE'] = rmse
    rmse_1=math.sqrt(sstot/(len(y)-1))
    results['rRMSE'] = rmse/rmse_1
    return results



data = pd.read_csv(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\Points\VPM_TROPOextend_1.csv",header=0,dtype=np.float64)
data1 = pd.read_csv(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\Points\VPM_TROPOextend_2.csv",header=0,dtype=np.float64)

data=np.array(data)
data1=np.array(data1)

m,n=data1.shape
pointID_1=data[:,1].tolist()
pointID_2=data1[:,1].tolist()
data_t=np.zeros((m,n))
for i in range(len(pointID_1)):
    index=pointID_1.index(pointID_2[i])
    data_t[index,:]=data1[i,:]
data1=data_t


kk=192
FID=data[:,0]
data_1=np.where(data[:,3:3+kk]>-5,data[:,3:3+kk],None)
data_2=np.where(data[:,3+kk:3+2*kk]>-5,data[:,3+kk:3+2*kk],None)
data_3=np.where(data[:,3+2*kk:3+3*kk]>0,data[:,3+2*kk:3+3*kk],None)
data_4=np.where(data[:,3+3*kk:3+4*kk]>-100,data[:,3+3*kk:3+4*kk],None)

data_5=np.where(data1[:,3+kk:3+2*kk]>=0,data1[:,3+kk:3+2*kk],None)
data_6=np.where(data1[:,3+2*kk:3+3*kk]>=0,data1[:,3+2*kk:3+3*kk],None)
data_7=np.where(data1[:,3+3*kk:3+4*kk]>=0,data1[:,3+3*kk:3+4*kk],None)

'''
PP1=11
PP2=0
PP3=1
data_1_1=data_1[:,PP1:kk:12]
data_1_2=data_1[:,PP2:kk:12]
data_1_3=data_1[:,PP3:kk:12]
data_2_1=data_2[:,PP1:kk:12]
data_2_2=data_2[:,PP2:kk:12]
data_2_3=data_2[:,PP3:kk:12]
data_3_1=data_3[:,PP1:kk:12]
data_3_2=data_3[:,PP2:kk:12]
data_3_3=data_3[:,PP3:kk:12]
data_4_1=data_4[:,PP1:kk:12]
data_4_2=data_4[:,PP2:kk:12]
data_4_3=data_4[:,PP3:kk:12]
data_5_1=data_5[:,PP1:kk:12]
data_5_2=data_5[:,PP2:kk:12]
data_5_3=data_5[:,PP3:kk:12]
data_6_1=data_6[:,PP1:kk:12]
data_6_2=data_6[:,PP2:kk:12]
data_6_3=data_6[:,PP3:kk:12]
data_7_1=data_7[:,PP1:kk:12]
data_7_2=data_7[:,PP2:kk:12]
data_7_3=data_7[:,PP3:kk:12]
data_1=np.hstack((data_1_1,data_1_2,data_1_3))
data_2=np.hstack((data_2_1,data_2_2,data_2_3))
data_3=np.hstack((data_3_1,data_3_2,data_3_3))
data_4=np.hstack((data_4_1,data_4_2,data_4_3))
data_5=np.hstack((data_5_1,data_5_2,data_5_3))
data_6=np.hstack((data_6_1,data_6_2,data_6_3))
data_7=np.hstack((data_7_1,data_7_2,data_7_3))
'''


listR_GPP = []
listR_SIF = []
listR_GPP_rRMSE = []
listR_SIF_rRMSE = []
listR_GPPcoff_PAR = []
listR_GPPcoff_TMP = []
listR_GPPcoff_NDVI = []
listR_GPPcoff_VPD = []
listR_GPPcoff_SM = []
listR_SIFcoff_PAR = []
listR_SIFcoff_TMP = []
listR_SIFcoff_NDVI = []
listR_SIFcoff_VPD = []
listR_SIFcoff_SM = []



for i in range (0,data_1.shape[0]):

    listGPP = data_1[i]
    listSIF = data_2[i]
    listPAR = data_3[i]
    listTMP = data_4[i]
    listVPD = data_5[i]
    listSM = data_6[i]
    listNDVI = data_7[i]

    list_GPP = []
    list_SIF = []
    list_PAR = []
    list_TMP = []
    list_VPD = []
    list_SM = []
    list_NDVI=[]
    for j in range(0,int(kk)):
        if listGPP[j]!=None and listSIF[j]!=None and listPAR[j]!=None and listTMP[j]!=None and listNDVI[j]!=None and listVPD[j]!=None and listSM[j]!=None:
            list_GPP.append(listGPP[j])
            list_SIF.append(listSIF[j])
            list_PAR.append(listPAR[j])
            list_TMP.append(listTMP[j])
            list_NDVI.append(listNDVI[j])
            list_VPD.append(listVPD[j])
            list_SM.append(listSM[j])
    print(i)
    x = {"list_PAR": list_PAR,"list_TMP": list_TMP,"list_VPD":list_VPD,"list_SM":list_SM,"list_NDVI": list_NDVI}
    y = {"list_GPP": list_GPP,"list_SIF": list_SIF}
    #y = {"list_GPP": list_GPP}
    DATAx = DataFrame(x)
    DATAy = DataFrame(y)

    if len(list_SIF)>10:
        try:
            X_train, X_test, y_train, y_test = train_test_split(DATAx, DATAy, test_size=0.2, random_state=0)
            sc = StandardScaler()
            X_train = sc.fit_transform(X_train)
            X_test = sc.fit_transform(X_test)

            #normalizer_x = preprocessing.scale(DATAx)
            #normalizer_y = preprocessing.scale(DATAy)
            #normalizer_x=normalizer_x.to_numpy()
            #normalizer_y = normalizer_y.to_numpy()
            pls = PLSRegression(n_components=3)
            pls.fit(X_train,y_train)
            coff=pls.coef_
            Ypredict = pls.predict(X_test)
            #R2Y = pls.score(normalizer_x,normalizer_y[:,0])
            #R, a = stats.pearsonr(normalizer_y[:,0], Ypredict[:,0])
            #R1, a1 = stats.pearsonr(normalizer_y[:,1], Ypredict[:,1])
            R_TOTAL = polyfit(y_test['list_GPP'], Ypredict[:,0], 1)
            R1_TOTAL = polyfit(y_test['list_SIF'], Ypredict[:, 1], 1)
            R= R_TOTAL['determination']
            R_rRMSE = R_TOTAL['rRMSE']
            R1 = R1_TOTAL['determination']
            R1_rRMSE = R1_TOTAL['rRMSE']
            #R = r2_score(normalizer_y[:,0], Ypredict[:,0])
            #R1 = r2_score(normalizer_y[:, 1], Ypredict[:, 1])

            [m,n]=np.shape(coff)
            for p in range(m):
                for q in range(n):
                    if coff[p][q]!=None:
                        if coff[p][q]>100 or coff[p][q]<-100:
                            coff[p][q] = None

            listR_GPP.append(R)
            listR_SIF.append(R1)
            listR_GPP_rRMSE.append(R_rRMSE)
            listR_SIF_rRMSE.append(R1_rRMSE)
            listR_GPPcoff_PAR.append(coff[0][0])
            listR_GPPcoff_TMP.append(coff[1][0])
            listR_GPPcoff_VPD.append(coff[2][0])
            listR_GPPcoff_SM.append(coff[3][0])
            listR_GPPcoff_NDVI.append(coff[4][0])
            listR_SIFcoff_PAR.append(coff[0][1])
            listR_SIFcoff_TMP.append(coff[1][1])
            listR_SIFcoff_VPD.append(coff[2][1])
            listR_SIFcoff_SM.append(coff[2][1])
            listR_SIFcoff_NDVI.append(coff[4][1])

        except:
            print("***********")
            rsquared = None
            listR_GPP.append(rsquared)
            listR_SIF.append(rsquared)
            listR_GPP_rRMSE.append(rsquared)
            listR_SIF_rRMSE.append(rsquared)
            listR_GPPcoff_PAR.append(rsquared)
            listR_GPPcoff_TMP.append(rsquared)
            listR_GPPcoff_VPD.append(rsquared)
            listR_GPPcoff_SM.append(rsquared)
            listR_GPPcoff_NDVI.append(rsquared)
            listR_SIFcoff_PAR.append(rsquared)
            listR_SIFcoff_TMP.append(rsquared)
            listR_SIFcoff_VPD.append(rsquared)
            listR_SIFcoff_SM.append(rsquared)
            listR_SIFcoff_NDVI.append(rsquared)

    else:
        rsquared =  None
        listR_GPP.append(rsquared)
        listR_SIF.append(rsquared)
        listR_GPP_rRMSE.append(rsquared)
        listR_SIF_rRMSE.append(rsquared)
        listR_GPPcoff_PAR.append(rsquared)
        listR_GPPcoff_TMP.append(rsquared)
        listR_GPPcoff_VPD.append(rsquared)
        listR_GPPcoff_SM.append(rsquared)
        listR_GPPcoff_NDVI.append(rsquared)
        listR_SIFcoff_PAR.append(rsquared)
        listR_SIFcoff_TMP.append(rsquared)
        listR_SIFcoff_VPD.append(rsquared)
        listR_SIFcoff_SM.append(rsquared)
        listR_SIFcoff_NDVI.append(rsquared)

book=openpyxl.Workbook()
sheet=book.create_sheet(index=0)
sheet.cell(column=1, row=1, value='FID')
sheet.cell(column=2, row=1, value='listR_GPP')
sheet.cell(column=3, row=1, value='listR_SIF')
sheet.cell(column=4, row=1, value='listR_GPP_rRMSE')
sheet.cell(column=5, row=1, value='listR_SIF_rRMSE')
sheet.cell(column=6, row=1, value='listR_GPPcoff_PAR')
sheet.cell(column=7, row=1, value='listR_GPPcoff_TMP')
sheet.cell(column=8, row=1, value='listR_GPPcoff_VPD')
sheet.cell(column=9, row=1, value='listR_GPPcoff_SM')
sheet.cell(column=10, row=1, value='listR_GPPcoff_NDVI')
sheet.cell(column=11, row=1, value='listR_SIFcoff_PAR')
sheet.cell(column=12, row=1, value='listR_SIFcoff_TMP')
sheet.cell(column=13, row=1, value='listR_SIFcoff_VPD')
sheet.cell(column=14, row=1, value='listR_SIFcoff_SM')
sheet.cell(column=15, row=1, value='listR_SIFcoff_NDVI')


for m in range(len(listR_GPP)):
    print(m)

    sheet.cell(column=1, row=m + 2, value=FID[m])
    sheet.cell(column=2, row=m + 2, value=listR_GPP[m])
    sheet.cell(column=3, row=m + 2, value=listR_SIF[m])
    sheet.cell(column=4, row=m + 2, value=listR_GPP_rRMSE[m])
    sheet.cell(column=5, row=m + 2, value=listR_SIF_rRMSE[m])
    sheet.cell(column=6, row=m + 2, value=listR_GPPcoff_PAR[m])
    sheet.cell(column=7, row=m + 2, value=listR_GPPcoff_TMP[m])
    sheet.cell(column=8, row=m + 2, value=listR_GPPcoff_VPD[m])
    sheet.cell(column=9, row=m + 2, value=listR_GPPcoff_SM[m])
    sheet.cell(column=10, row=m + 2, value=listR_GPPcoff_NDVI[m])
    sheet.cell(column=11, row=m + 2, value=listR_SIFcoff_PAR[m])
    sheet.cell(column=12, row=m + 2, value=listR_SIFcoff_TMP[m])
    sheet.cell(column=13, row=m + 2, value=listR_SIFcoff_VPD[m])
    sheet.cell(column=14, row=m + 2, value=listR_SIFcoff_SM[m])
    sheet.cell(column=15, row=m + 2, value=listR_SIFcoff_NDVI[m])


book.save(r'E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\PLSR\VPM_TROPOextend.xlsx')
print()

