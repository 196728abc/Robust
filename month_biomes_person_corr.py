
import glob
import numpy as np
import pandas as pd
from sklearn.metrics import r2_score
from pandas.core.frame import DataFrame
import os
import xlwt
from collections import Counter
import pingouin as pg
import matplotlib.pyplot as plt
import openpyxl
import math
import statsmodels.api as sm
from sklearn import preprocessing

def min_max(List):
    min_1=np.min(np.array(List))
    max_1=np.max(np.array(List))
    result=(np.array(List)-min_1)/(max_1-min_1)
    return result


def polyfit(x,y,degree):
    results = {}
    coeffs = np.polyfit(x,y,degree)
    results['polynomial'] = coeffs.tolist()
    p = np.poly1d(coeffs)
    yhat = p(x)
    ybar = np.sum(y)/len(y)
    ssreg = np.sum((yhat - ybar)**2)
    sstot = np.sum((y - ybar)**2)
    results['determination'] = ssreg/sstot
    return results


book=openpyxl.Workbook()
sheet=book.create_sheet('R2',index=0)
for i in range(1,13):
    sheet.cell(column=i, row=1, value=i)

#data_kinds=['FLUXCOM_CSIF','FLUXCOM_GOMESAT','FLUXCOM_TROPOextend','PML_CSIF','PML_GOMESAT','PML_TROPOextend','VPM_CSIF','VPM_GOMESAT','VPM_TROPOextend']
#data_len=[192,192,216,168,180,180,192,168,192]
data_kinds=['VPM_CSIF']
data_len=[192]

aaaa=[221,222,223,224]

for jj in range(0,1):
    data1 = pd.read_csv('C:/Users/lijia/Desktop/TEST/relationship bwteen R2 and NDVI/Points_extract by North landcover/ENF/'+data_kinds[jj]+'_1.csv',header=0,dtype=np.float64)
    data2 = pd.read_csv('C:/Users/lijia/Desktop/TEST/relationship bwteen R2 and NDVI/Points_extract by North landcover/ENF/'+data_kinds[jj]+'_2.csv',header=0,dtype=np.float64)

    data1=np.array(data1)
    data2=np.array(data2)

    m,n=data2.shape
    pointID_1=data1[:,0].tolist()
    pointID_2=data2[:,0].tolist()
    data_t=np.zeros((m,n))
    for i in range(len(pointID_1)):
        index=pointID_1.index(pointID_2[i])
        data_t[index,:]=data2[i,:]
    data2=data_t

    kk=data_len[jj]
    FID=data1[:,0]
    data_1_gpp=np.where(data1[:,3:3+int(kk/2)]>-5,data1[:,3:3+int(kk/2)],None)
    data_1_sif=np.where(data1[:,3+int(kk/2):3+kk]>-5,data1[:,3+int(kk/2):3+kk],None)
    data_2_gpp=np.where(data2[:,3:3+int(kk/2)]>-5,data2[:,3:3+int(kk/2)],None)
    data_2_sif=np.where(data2[:,3+int(kk/2):3+kk]>-5,data2[:,3+int(kk/2):3+kk],None)

    data_1 = np.hstack((data_1_gpp, data_2_gpp))
    data_2 = np.hstack((data_1_sif, data_2_sif))

    data_GPP=np.zeros((int(kk/4)*data_1.shape[0],4))
    data_SIF=np.zeros((int(kk/4)*data_2.shape[0],4))
    iii=0
    for i in range(0,12,3):
        if i==0:
            pp1_1 = data_1[:, 0:kk:12]
            pp1_2 = data_1[:, 1:kk:12]
            pp1_3 = data_1[:, 11:kk:12]
            pp2_1 = data_2[:, 0:kk:12]
            pp2_2 = data_2[:, 1:kk:12]
            pp2_3 = data_2[:, 11:kk:12]
        else:
            pp1_1 = data_1[:, i-1:kk:12]
            pp1_2 = data_1[:, i:kk:12]
            pp1_3 = data_1[:, i+1:kk:12]
            pp2_1 = data_2[:, i-1:kk:12]
            pp2_2 = data_2[:, i:kk:12]
            pp2_3 = data_2[:, i+1:kk:12]

        pp1=np.hstack((pp1_1, pp1_2,pp1_3))
        pp2 = np.hstack((pp2_1, pp2_2, pp2_3))
        data_GPP[:,iii]=np.reshape(pp1,(int(kk/4)*data_1.shape[0],))
        data_SIF[:,iii]=np.reshape(pp2,(int(kk/4)*data_1.shape[0],))
        iii=iii+1
    listR = []
    listC=[]
    None_list=np.zeros(4,)
    for i in range (0,4):
        listGPP = data_GPP[:,i]
        listSIF = data_SIF[:,i]
        list_GPP=[]
        list_SIF=[]
        for j in range(0,int(kk/4)*data_1.shape[0]):
            if np.isnan(listGPP[j]) or np.isnan(listSIF[j]):
                print()
                None_list[i]=None_list[i]+1
            else:
                list_GPP.append(listGPP[j])
                list_SIF.append(listSIF[j])
        print(i)

        plt.subplot(aaaa[i])
        plt.scatter(list_SIF,list_GPP)


        if len(list_GPP)>4:
            try:
                R1 = polyfit(list_SIF, list_GPP, 1)['determination']
                listR.append(R1)
            except:
                rsquared = None
                listR.append(rsquared)
        else:
            rsquared = None
            listR.append(rsquared)
        for m in range(len(listR)):
            sheet.cell(column=m + 1, row=jj + 2, value=listR[m])
plt.show()
book.save(r'C:\Users\lijia\Desktop\TEST\R2_GPP_SIF_Month.xlsx')
print()
