
import glob
import numpy as np
import pandas as pd
from sklearn.metrics import r2_score
from pandas.core.frame import DataFrame
import os
import xlwt
from collections import Counter
import pingouin as pg
import matplotlib.pyplot as plt
import openpyxl
import math



book=openpyxl.Workbook()
sheet=book.create_sheet(index=0)
for i in range(1,13):
    sheet.cell(column=i, row=1, value=i)

data_kinds=['FLUXCOM_CSIF','FLUXCOM_GOMESAT','FLUXCOM_TROPOextend']
data_len=[192,192,216]
for jj in range(0,3):

    data = pd.read_csv('C:/Users/lijia/Desktop/TEST/relationship bwteen R2 and NDVI/biome types/GRA/'+data_kinds[jj]+'.csv',header=0,dtype=np.float64)
    #data1 = pd.read_csv(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\CRU\Point_GPP_SIF_meteo\VPM_GOSIF_2.csv",header=0,dtype=np.float64)

    data=np.array(data)


    kk=data_len[jj]
    FID=data[:,0]
    #data_1=np.where(data[:,3:3+kk]>-5,data[:,3:3+kk],None)
    #data_2=np.where(data[:,3+kk:3+2*kk]>-5,data[:,3+kk:3+2*kk],None)
    data_1=np.where(data[:,3+2*kk:3+3*kk]>-5,data[:,3+2*kk:3+3*kk],None)


    data_NDVI=np.zeros((int(kk/12)*data_1.shape[0],12))
    for i in range(0,12):
        pp1=data_1[:,i:kk:12]
        data_NDVI[:,i]=np.reshape(pp1,(int(kk/12)*data_1.shape[0],))

    listR = []
    for i in range (0,12):
        listNDVI = data_NDVI[:,i]
        list_NDVI=[]
        for j in range(0,int(kk/12)*data_1.shape[0]):
            if np.isnan(listNDVI[j]):
                print()
            else:
                list_NDVI.append(listNDVI[j])
        dd=np.mean(list_NDVI)
        listR.append(dd)

    for m in range(len(listR)):
        sheet.cell(column=m+1, row=jj+2, value=listR[m])

book.save(r'C:\Users\lijia\Desktop\TEST\relationship bwteen R2 and NDVI\biome types\GRA\NDVI.xlsx')
print()
