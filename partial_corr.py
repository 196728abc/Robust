
import glob
import numpy as np
import pandas as pd
from pandas.core.frame import DataFrame
import os
import xlwt
from collections import Counter
import pingouin as pg
from sklearn.linear_model import LinearRegression
import openpyxl as op
import scipy.stats as stats
from matplotlib.font_manager import FontProperties
from statsmodels.formula.api import ols
#from sklearn.preprocessing import StandardScaler
import statsmodels.api as sm
import openpyxl



def standardize(x):
   # return (x - np.mean(x))/(np.std(x) )
    return (x - np.min(x))/(np.max(x)-np.min(x))

data = pd.read_csv(r"C:\Users\lijia\Desktop\TEST\Points\VPM_TROPOextend_1.csv",header=0,dtype=np.float64)
data1 = pd.read_csv(r"C:\Users\lijia\Desktop\TEST\Points\VPM_TROPOextend_2.csv",header=0,dtype=np.float64)

data=np.array(data)
data1=np.array(data1)

m,n=data1.shape
pointID_1=data[:,1].tolist()
pointID_2=data1[:,1].tolist()
data_t=np.zeros((m,n))
for i in range(len(pointID_1)):
    index=pointID_1.index(pointID_2[i])
    data_t[index,:]=data1[i,:]
data1=data_t

#Point1=data[:,1]
#Point2=data1[:,1]
#ddd=np.max(abs(Point1-Point2))

kk=192
FID=data[:,0]
data_0=np.where(data[:,3:3+kk]>-5,data[:,3:3+kk],None)
data_1=np.where(data[:,3+kk:3+2*kk]>-5,data[:,3+kk:3+2*kk],None)
data_2=np.where(data[:,3+2*kk:3+3*kk]>0,data[:,3+2*kk:3+3*kk],None)
data_3=np.where(data[:,3+3*kk:3+4*kk]>-100,data[:,3+3*kk:3+4*kk],None)

#data_1=np.where(data[:,3:3+kk]>0,data[:,3:3+kk],None)
data_4=np.where(data1[:,3+kk:3+2*kk]>=0,data1[:,3+kk:3+2*kk],None)
data_5=np.where(data1[:,3+2*kk:3+3*kk]>=0,data1[:,3+2*kk:3+3*kk],None)
#data_6=np.where(data1[:,3+3*kk:]>0,data1[:,3+3*kk:],None)


'''
data_T_1=np.append(data_1,data_2,axis=1)
data_T_1=np.append(data_T_1,data_3,axis=1)
data_T_2=np.append(data_4,data_5,axis=1)
data_T_2=np.append(data_T_2,data_6,axis=1)
'''

listR_PAR = []
listR_TMP = []
listR_PRE = []
#listR_VPD = []
#listR_SM = []

for i in range (0,data_1.shape[0]):
    listGPP = data_0[i]
    #listSIF = data_1[i]
    listPAR = data_2[i]
    listTMP = data_3[i]
    listPRE = data_4[i]
    listVPD = data_5[i]
    #listSM = data_6[i]

    list_GPP=[]
    list_PAR=[]
    list_TMP = []
    list_PRE = []
    #list_VPD = []
    #list_SM = []
    for j in range(0,kk):
        if listGPP[j]!=None and listPAR[j]!=None and listTMP[j]!=None and listPRE[j]!=None:# and listPRE[j]!=None and listSM[j]!=None:
            list_GPP.append(listGPP[j])
            list_PAR.append(listPAR[j])
            list_TMP.append(listTMP[j])
            list_PRE.append(listPRE[j])
            #list_VPD.append(listVPD[j])
            #list_SM.append(listSM[j])
    print(i)
    c = {"list_GPP": list_GPP, "list_PAR":list_PAR,"list_TMP": list_TMP,"list_PRE":list_PRE}#, "list_PRE":list_PRE,"list_SM": list_SM}
    DATAF = DataFrame(c)

    if len(list_GPP)>5:
        #r,p = stats.pearsonr(list_x, list_y)
        try:
            AA=pg.partial_corr(data=DATAF, x='list_PAR', y='list_GPP', covar=['list_TMP','list_PRE'],method='spearman').round(3)
            rsquared = AA.iloc[0].at['r']
            p=AA.iloc[0].at['p-val']
            if p > 0.05:
                listR_PAR.append(0)
            else:
                listR_PAR.append(rsquared)

            AA=pg.partial_corr(data=DATAF, x='list_TMP', y='list_GPP', covar=['list_PAR','list_PRE'],method='spearman').round(3)
            rsquared = AA.iloc[0].at['r']
            p=AA.iloc[0].at['p-val']
            if p > 0.05:
                listR_TMP.append(0)
            else:
                listR_TMP.append(rsquared)

            AA=pg.partial_corr(data=DATAF, x='list_PRE', y='list_GPP', covar=['list_PAR', 'list_TMP'],method='spearman').round(3)
            rsquared = AA.iloc[0].at['r']
            p=AA.iloc[0].at['p-val']
            if p > 0.05:
                listR_PRE.append(0)
            else:
                listR_PRE.append(rsquared)
            '''
            AA=pg.partial_corr(data=DATAF, x='list_VPD', y='list_GPP', covar=['list_PAR','list_TMP','list_PRE', 'list_SM'],method='spearman').round(3)
            rsquared = AA.iloc[0].at['r']
            p=AA.iloc[0].at['p-val']
            if p > 0.05:
                listR_VPD.append(0)
            else:
                listR_VPD.append(rsquared)

            AA=pg.partial_corr(data=DATAF, x='list_SM', y='list_GPP', covar=['list_PAR','list_TMP','list_PRE','list_VPD'],method='spearman').round(3)
            rsquared = AA.iloc[0].at['r']
            p=AA.iloc[0].at['p-val']
            if p > 0.05:
                listR_SM.append(0)
            else:
                listR_SM.append(rsquared)
            '''
        except:
            rsquared = None
            listR_PAR.append(rsquared)
            listR_TMP.append(rsquared)
            listR_PRE.append(rsquared)
            #listR_VPD.append(rsquared)
            #listR_SM.append(rsquared)

    else:
        rsquared = None
        listR_PAR.append(rsquared)
        listR_TMP.append(rsquared)
        listR_PRE.append(rsquared)
        #listR_VPD.append(rsquared)
        #listR_SM.append(rsquared)

book=openpyxl.Workbook()
sheet=book.create_sheet(index=0)
sheet.cell(column=1, row=1, value='FID')
sheet.cell(column=2, row=1, value='listR_PAR')
sheet.cell(column=3, row=1, value='listR_TMP')
sheet.cell(column=4, row=1, value='listR_PRE')
#sheet.cell(column=5, row=1, value='listR_VPD')
#sheet.cell(column=6, row=1, value='listR_SM')

for m in range(len(listR_TMP)):
    print(m)
    sheet.cell(column=1, row=m+2, value=FID[m])
    sheet.cell(column=2, row=m+2, value=listR_PAR[m])
    sheet.cell(column=3, row=m+2, value=listR_TMP[m])
    sheet.cell(column=4, row=m+2, value=listR_PRE[m])
    #sheet.cell(column=5, row=m+2, value=listR_VPD[m])
    #sheet.cell(column=6, row=m+2, value=listR_SM[m])

book.save(r'C:\Users\lijia\Desktop\TEST\partial corr\VPM_METEO_from_TROPOextend.xlsx')
print()
