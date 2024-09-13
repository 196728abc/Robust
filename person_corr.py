
import glob
import math

import numpy as np
import pandas as pd
from sklearn.metrics import r2_score
from pandas.core.frame import DataFrame
import os
import xlwt
from collections import Counter
import openpyxl
import scipy.stats as stats

'''
def polyfit(x,y,degree):
    results = {}
    coeffs = np.polyfit(x,y,degree)
    results['polynomial'] = coeffs.tolist()
    p = np.poly1d(coeffs)
    yhat = p(x)
    ybar = np.sum(y)/len(y)
    ssreg = np.sum((yhat - ybar)**2)
    sstot = np.sum((y - ybar)**2)
    results['determination'] = ssreg/sstot

    rmse=math.sqrt(np.sum((y-yhat)**2)/len(y))
    results['RMSE'] = rmse
    rmse_1=math.sqrt(sstot/(len(y)-1))
    results['rRMSE'] = rmse/rmse_1
    return results
'''

data = pd.read_csv(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\GPP_SIF\GPP_GPP or SIF_SIF\GPP_GPP\VPM_PML.csv",header=0,dtype=np.float64)
#data1 = pd.read_csv(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\Points\FLUXCOM_CSIF_2.csv",header=0,dtype=np.float64)
data=np.array(data)
'''
data1=np.array(data1)
m,n=data1.shape
pointID_1=data[:,1].tolist()
pointID_2=data1[:,1].tolist()
data_t=np.zeros((m,n))
for i in range(len(pointID_1)):
    index=pointID_1.index(pointID_2[i])
    data_t[index,:]=data1[i,:]
data1=data_t
'''

kk=168
FID=data[:,0]
#data_1=np.where(data[:,3+6*12+3:3+kk-54]>-5,data[:,3+6*12+3:3+kk-54],None)
#data_2=np.where(data[:,3+kk+6*12+3:3+2*kk-54]>-5,data[:,3+kk+6*12+3:3+2*kk-54],None)
data_1=np.where(data[:,3:3+kk]>0,data[:,3:3+kk],None)
data_2=np.where(data[:,3+kk:3+2*kk]>0,data[:,3+kk:3+2*kk],None)

#data_2=np.where(data1[:,3+2*kk:3+3*kk]>0,data1[:,3+2*kk:3+3*kk],None)
#data_2=np.where(data1[:,3+3*kk:3+4*kk]>0,data1[:,3+3*kk:3+4*kk],None)
'''
data_1_1=data_1[:,5:kk:12]
data_1_2=data_1[:,6:kk:12]
data_1_3=data_1[:,7:kk:12]
data_2_1=data_2[:,5:kk:12]
data_2_2=data_2[:,6:kk:12]
data_2_3=data_2[:,7:kk:12]
data_1=np.hstack((data_1_1,data_1_2,data_1_3))
data_2=np.hstack((data_2_1,data_2_2,data_2_3))
'''

listR2=[]
RMSE=[]
rRMSE=[]

#listR_SM = []

for i in range (0,data_1.shape[0]):

    listGPP = data_1[i]
    listSIF = data_2[i]
    #listNDVI = data_3[i]
    #listTMP = data_4[i]
    #listPRE = data_5[i]
    #listVPD = data_6[i]
    #listSM = data_7[i]

    list_GPP = []
    list_SIF = []
    #list_NDVI = []
    #list_TMP = []
    #list_PRE = []
    #list_VPD = []
    #list_SM = []
    for j in range(0,int(kk)):
        if listGPP[j]!=None and listSIF[j]!=None: # and listNDVI[j]!=None:# and listTMP[j]!=None and listPRE[j]!=None:# and listVPD[j]!=None:# and listSM[j]!=None:
            list_GPP.append(listGPP[j])
            list_SIF.append(listSIF[j])
            #list_SW.append(listSW[j])
            #list_TMP.append(listTMP[j])
            #list_PRE.append(listPRE[j])
            #list_VPD.append(listVPD[j])
            #list_SM.append(listSM[j])
    print(i)
    if len(list_GPP)>4:
        try:
            #R1 = np.corrcoef(list_GPP, list_SIF)
            #listR_SIF.append(R1[0][1])
            #R1 = np.corrcoef(list_GPP, list_SW)
            #listR_SW.append(R1[0][1])
            #R2 = np.corrcoef(list_GPP, list_TMP)
            #listR_TMP.append(R2[0][1])
            #R3 = np.corrcoef(list_GPP, list_PRE)
            #listR_PRE.append(R3[0][1])
            #R4 = np.corrcoef(list_GPP, list_VPD)
            #listR_VPD.append(R4[0][1])

            R, a = stats.pearsonr(list_GPP, list_SIF)
            listR2.append(R)
            #R1 = polyfit(list_GPP, list_SIF, 1)
            #listR2.append(R1['determination'])
            #RMSE.append(R1['RMSE'])
            #rRMSE.append(R1['rRMSE'])

            #R1 = np.corrcoef(list_GPP, list_SIF)  ##########
            #listR2.append(R1[0][1])


            #R2 = polyfit(list_GPP, list_SW, 1)['determination']
            #listR_SW.append(R2)
            #R3 = polyfit(list_GPP, list_TMP, 1)['determination']
            #listR_TMP.append(R3)
            #R4 = polyfit(list_GPP, list_PRE, 1)['determination']
            #listR_PRE.append(R4)
            '''
            R4 = polyfit(list_GPP, list_VPD, 1)['determination']
            listR_VPD.append(R4)
            #R5 = polyfit(list_GPP, list_SM, 1)['determination']
            #listR_SM.append(R5)
            '''
        except:
            rsquared = None
            listR2.append(rsquared)

            #listR_SW.append(rsquared)
            #listR_TMP.append(rsquared)
            #listR_PRE.append(rsquared)
            #listR_VPD.append(rsquared)
            #listR_SM.append(rsquared)
    else:
        rsquared =  None
        listR2.append(rsquared)

        #listR_SW.append(rsquared)
        #listR_TMP.append(rsquared)
        #listR_PRE.append(rsquared)
        #listR_VPD.append(rsquared)
        #listR_SM.append(rsquared)

book=openpyxl.Workbook()
sheet=book.create_sheet(index=0)
sheet.cell(column=1, row=1, value='FID')
sheet.cell(column=2, row=1, value='listR2')

#sheet.cell(column=5, row=1, value='listR_NIRv')
#sheet.cell(column=5, row=1, value='listR_VPD')
#sheet.cell(column=6, row=1, value='listR_SM')


for m in range(len(listR2)):
    print(m)
    sheet.cell(column=1, row=m + 2, value=FID[m])
    sheet.cell(column=2, row=m + 2, value=listR2[m])

    #sheet.cell(column=5, row=m + 2, value=listR_PRE[m])
    #sheet.cell(column=5, row=m + 2, value=listR_VPD[m])
    #sheet.cell(column=6, row=m + 2, value=listR_SM[m])

book.save(r'E:/1_canopy_photosynthesis/photosynthesis_and_Global_Climate/SIF/data/GPP_SIF/GPP_GPP or SIF_SIF/GPP_GPP/R_VPM_PML.xlsx')
print()
