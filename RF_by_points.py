import glob
import numpy as np
import pandas as pd
import math
from pandas.core.frame import DataFrame
import pingouin as pg
from sklearn.linear_model import LinearRegression
import openpyxl as op
from sklearn import preprocessing
import openpyxl
from sklearn.cross_decomposition import PLSRegression
from sklearn.metrics import r2_score
from sklearn.preprocessing import StandardScaler
from sklearn.cross_decomposition import PLSRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split

def polyfit(x,y,degree):
    results = {}
    coeffs = np.polyfit(x,y,degree)
    results['polynomial'] = coeffs.tolist()
    p = np.poly1d(coeffs)
    yhat = p(x)
    ybar = np.sum(y)/len(y)
    ssreg = np.sum((yhat - ybar)**2)
    sstot = np.sum((y - ybar)**2)
    results['determination'] = ssreg/sstot

    rmse=math.sqrt(np.sum((y-yhat)**2)/len(y))
    results['RMSE'] = rmse
    rmse_1=math.sqrt(sstot/(len(y)-1))
    results['rRMSE'] = rmse/rmse_1
    return results


data = pd.read_csv(
    r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\Points\PML_CSIF_1.csv",
    header=0, dtype=np.float64)
data1 = pd.read_csv(
    r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\Points\PML_CSIF_2.csv",
    header=0, dtype=np.float64)

data = np.array(data)
data1 = np.array(data1)

m, n = data1.shape
pointID_1 = data[:, 1].tolist()
pointID_2 = data1[:, 1].tolist()
data_t = np.zeros((m, n))
for i in range(len(pointID_1)):
    index = pointID_1.index(pointID_2[i])
    data_t[index, :] = data1[i, :]
data1 = data_t

kk = 168
FID = data[:, 0]
data_1 = np.where(data[:, 3:3 + kk] > -5, data[:, 3:3 + kk], None)
data_2 = np.where(data[:, 3 + kk:3 + 2 * kk] > -5, data[:, 3 + kk:3 + 2 * kk], None)
data_3 = np.where(data[:, 3 + 2 * kk:3 + 3 * kk] > 0, data[:, 3 + 2 * kk:3 + 3 * kk], None)
data_4 = np.where(data[:, 3 + 3 * kk:3 + 4 * kk] > -100, data[:, 3 + 3 * kk:3 + 4 * kk], None)

data_5 = np.where(data1[:, 3 + kk:3 + 2 * kk] >= 0, data1[:, 3 + kk:3 + 2 * kk], None)
data_6 = np.where(data1[:, 3 + 2 * kk:3 + 3 * kk] >= 0, data1[:, 3 + 2 * kk:3 + 3 * kk], None)
data_7 = np.where(data1[:, 3 + 3 * kk:3 + 4 * kk] >= 0, data1[:, 3 + 3 * kk:3 + 4 * kk], None)
'''
PP1 = 11
PP2 = 0
PP3 = 1
data_1_1 = data_1[:, PP1:kk:12]
data_1_2 = data_1[:, PP2:kk:12]
data_1_3 = data_1[:, PP3:kk:12]
data_2_1 = data_2[:, PP1:kk:12]
data_2_2 = data_2[:, PP2:kk:12]
data_2_3 = data_2[:, PP3:kk:12]
data_3_1 = data_3[:, PP1:kk:12]
data_3_2 = data_3[:, PP2:kk:12]
data_3_3 = data_3[:, PP3:kk:12]
data_4_1 = data_4[:, PP1:kk:12]
data_4_2 = data_4[:, PP2:kk:12]
data_4_3 = data_4[:, PP3:kk:12]
data_5_1 = data_5[:, PP1:kk:12]
data_5_2 = data_5[:, PP2:kk:12]
data_5_3 = data_5[:, PP3:kk:12]
data_6_1 = data_6[:, PP1:kk:12]
data_6_2 = data_6[:, PP2:kk:12]
data_6_3 = data_6[:, PP3:kk:12]
data_7_1 = data_7[:, PP1:kk:12]
data_7_2 = data_7[:, PP2:kk:12]
data_7_3 = data_7[:, PP3:kk:12]
data_1 = np.hstack((data_1_1, data_1_2, data_1_3))
data_2 = np.hstack((data_2_1, data_2_2, data_2_3))
data_3 = np.hstack((data_3_1, data_3_2, data_3_3))
data_4 = np.hstack((data_4_1, data_4_2, data_4_3))
data_5 = np.hstack((data_5_1, data_5_2, data_5_3))
data_6 = np.hstack((data_6_1, data_6_2, data_6_3))
data_7 = np.hstack((data_7_1, data_7_2, data_7_3))
'''
listR_GPP = []
listR_SIF = []
listR_GPP_rRMSE = []
listR_SIF_rRMSE = []

for i in range(0, data_1.shape[0]):

    listGPP = data_1[i]
    listSIF = data_2[i]
    listPAR = data_3[i]
    listTMP = data_4[i]
    listVPD = data_5[i]
    listSM = data_6[i]
    listNDVI = data_7[i]

    list_GPP = []
    list_SIF = []
    list_PAR = []
    list_TMP = []
    list_NDVI = []
    list_VPD = []
    list_SM = []
    for j in range(0, int(kk)):
        if listGPP[j] != None and listSIF[j] != None and listPAR[j] != None and listTMP[j] != None and listNDVI[j] != None and listVPD[j] != None and listSM[j] != None:
            list_GPP.append(listGPP[j])
            list_SIF.append(listSIF[j])
            list_PAR.append(listPAR[j])
            list_TMP.append(listTMP[j])
            list_NDVI.append(listNDVI[j])
            list_VPD.append(listVPD[j])
            list_SM.append(listSM[j])
    print(i)
    x = {"list_PAR": list_PAR, "list_TMP": list_TMP, "list_VPD": list_VPD, "list_SM": list_SM,"list_NDVI": list_NDVI}
    y = {"list_GPP": list_GPP, "list_SIF": list_SIF}
    # y = {"list_GPP": list_GPP}
    DATAx = DataFrame(x)
    DATAy = DataFrame(y)

    if len(list_SIF) > 10:
        try:
            X_train, X_test, y_train, y_test = train_test_split(DATAx, DATAy, test_size=0.2, random_state=0)
            sc = StandardScaler()
            X_train = sc.fit_transform(X_train)
            X_test = sc.fit_transform(X_test)
            #X_train = sc.fit_transform(normalizer_x)
            #X_test = sc.fit_transform(X_test)

            regr = RandomForestRegressor(n_estimators=20, random_state=0)
            regr.fit(X_train, y_train)
            # score=regr.score(X_test,y_test)
            Ypredict = regr.predict(X_test)
            R_TOTAL = polyfit(y_test['list_GPP'], Ypredict[:, 0], 1)
            R1_TOTAL = polyfit(y_test['list_SIF'], Ypredict[:, 1], 1)
            R = R_TOTAL['determination']
            R_rRMSE = R_TOTAL['rRMSE']
            R1 = R1_TOTAL['determination']
            R1_rRMSE = R1_TOTAL['rRMSE']

            listR_GPP.append(R)
            listR_SIF.append(R1)
            listR_GPP_rRMSE.append(R_rRMSE)
            listR_SIF_rRMSE.append(R1_rRMSE)

        except:
            print("***********")
            rsquared = None
            listR_GPP.append(rsquared)
            listR_SIF.append(rsquared)
            listR_GPP_rRMSE.append(rsquared)
            listR_SIF_rRMSE.append(rsquared)

    else:
        rsquared = None
        listR_GPP.append(rsquared)
        listR_SIF.append(rsquared)
        listR_GPP_rRMSE.append(rsquared)
        listR_SIF_rRMSE.append(rsquared)

book = openpyxl.Workbook()
sheet = book.create_sheet(index=0)
sheet.cell(column=1, row=1, value='FID')
sheet.cell(column=2, row=1, value='listR_GPP')
sheet.cell(column=3, row=1, value='listR_SIF')
sheet.cell(column=4, row=1, value='listR_GPP_rRMSE')
sheet.cell(column=5, row=1, value='listR_SIF_rRMSE')

for m in range(len(listR_GPP)):
    print(m)
    sheet.cell(column=1, row=m + 2, value=FID[m])
    sheet.cell(column=2, row=m + 2, value=listR_GPP[m])
    sheet.cell(column=3, row=m + 2, value=listR_SIF[m])
    sheet.cell(column=4, row=m + 2, value=listR_GPP_rRMSE[m])
    sheet.cell(column=5, row=m + 2, value=listR_SIF_rRMSE[m])

book.save(
    r'E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\RF\PML_CSIF.xlsx')
print()

