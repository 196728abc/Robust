
import glob
import numpy as np
import pandas as pd
from pandas.core.frame import DataFrame
import os
import xlwt
from collections import Counter
import pingouin as pg
from sklearn.linear_model import LinearRegression
import openpyxl as op
import scipy.stats as stats
from matplotlib.font_manager import FontProperties
from statsmodels.formula.api import ols
from sklearn import preprocessing
from sklearn.decomposition import PCA
import statsmodels.api as sm
import openpyxl


data = pd.read_csv(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\CRU\Point_GPP_SIF_meteo\VPM_GOME.csv",header=0,dtype=np.float64)
#data1 = pd.read_csv(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\CRU\Point_GPP_SIF_meteo\MODIS_GOSIF_2.csv",header=0,dtype=np.float64)

data=np.array(data)
#data1=np.array(data1)
'''
m,n=data1.shape
pointID_1=data[:,1].tolist()
pointID_2=data1[:,1].tolist()
data_t=np.zeros((m,n))
for i in range(len(pointID_1)):
    index=pointID_1.index(pointID_2[i])
    data_t[index,:]=data1[i,:]
data1=data_t
'''

kk=120
FID=data[:,0]
data_1=np.where(data[:,3:3+kk]>-5,data[:,3:3+kk],None)
data_2=np.where(data[:,3+kk:3+2*kk]>-5,data[:,3+kk:3+2*kk],None)
data_3=np.where(data[:,3+2*kk:3+3*kk]>0,data[:,3+2*kk:3+3*kk],None)
data_4=np.where(data[:,3+3*kk:3+4*kk]>-100,data[:,3+3*kk:3+4*kk],None)

data_5=np.where(data[:,3+4*kk:3+5*kk]>=0,data[:,3+4*kk:3+5*kk],None)
data_6=np.where(data[:,3+5*kk:3+6*kk]>0,data[:,3+5*kk:3+6*kk],None)
data_7=np.where(data[:,3+6*kk:]>0,data[:,3+6*kk:],None)


listR_GPP_PAR = []
listR_GPP_TMP = []
listR_GPP_PRE = []
listR_GPP_VPD = []
listR_GPP_SM = []

listR_SIF_PAR = []
listR_SIF_TMP = []
listR_SIF_PRE = []
listR_SIF_VPD = []
listR_SIF_SM = []
listR_GPP = []
listR_SIF = []
listR_PAR = []
listR_TMP = []
listR_PRE = []
listR_VPD = []
listR_SM = []
listR = []
listR11=[]

for i in range (0,data_1.shape[0]):

    listGPP = data_1[i]
    listSIF = data_2[i]
    listPAR = data_3[i]
    listTMP = data_4[i]
    listPRE = data_5[i]
    listVPD = data_6[i]
    listSM = data_7[i]

    list_GPP=[]
    list_SIF = []
    list_PAR=[]
    list_TMP = []
    list_PRE = []
    list_VPD = []
    list_SM = []
    for j in range(0,kk):
        if listGPP[j]!=None and listPAR[j]!=None and listTMP[j]!=None and listSIF[j]!=None and listVPD[j]!=None and listSM[j]!=None and listPRE[j]!=None :
            list_GPP.append(listGPP[j])
            list_SIF.append(listSIF[j])
            list_PAR.append(listPAR[j])
            list_TMP.append(listTMP[j])
            list_PRE.append(listPRE[j])
            list_VPD.append(listVPD[j])
            list_SM.append(listSM[j])
    print(i)
    x = {"list_PAR": list_PAR,"list_TMP": list_TMP,"list_PRE": list_PRE,"list_VPD":list_VPD,"list_SM": list_SM}
    y = {"list_GPP": list_GPP, "list_SIF": list_SIF}
    DATAx = DataFrame(x)  # 将字典转换成为数据框
    DATAy = DataFrame(y)

    if len(list_GPP)>5:
        try:
            r = np.zeros((2,5))

            r[0, 0], a = stats.pearsonr(list_GPP, list_PAR)
            if a>0.05:
                r[0, 0]=0
            r[0, 1], a = stats.pearsonr(list_GPP, list_TMP)
            if a>0.05:
                r[0, 1]=0
            r[0, 2], a = stats.pearsonr(list_GPP, list_PRE)
            if a>0.05:
                r[0, 2]=0
            r[0, 3], a = stats.pearsonr(list_GPP, list_VPD)
            if a>0.05:
                r[0, 3]=0
            r[0, 4], a = stats.pearsonr(list_GPP, list_SM)
            if a>0.05:
                r[0, 4]=0
            r[1, 0], a = stats.pearsonr(list_SIF, list_PAR)
            if a>0.05:
                r[1, 0]=0
            r[1, 1], a = stats.pearsonr(list_SIF, list_TMP)
            if a>0.05:
                r[1, 1]=0
            r[1, 2], a = stats.pearsonr(list_SIF, list_PRE)
            if a>0.05:
                r[1, 2]=0
            r[1, 3], a = stats.pearsonr(list_SIF, list_VPD)
            if a>0.05:
                r[1, 3]=0
            r[1, 4], a = stats.pearsonr(list_SIF, list_SM)
            if a>0.05:
                r[1, 4]=0
            normalizer_x = preprocessing.scale(DATAx)
            # 沿着某个轴标准化数据集，以均值为中心，以分量为单位方差
            DATAx_normalized_x = pd.DataFrame(normalizer_x, columns=["list_PAR","list_TMP","list_PRE","list_VPD","list_SM"])
            #r,p = stats.pearsonr(list_x, list_y)
            pca = PCA()
            pca.fit(DATAx_normalized_x)  # 训练模型
            vector = pca.components_  # 返回模型的特征向量
           # vector_value = pca.explained_variance_  # 返回模型的特征值
            b_x = pca.explained_variance_ratio_  # 返回各成分的方差百分比
            C_x = np.dot(b_x,vector)
            F_x = np.sum( C_x * np.array(DATAx_normalized_x),axis=1)

            normalizer_y = preprocessing.scale(DATAy)
            # 沿着某个轴标准化数据集，以均值为中心，以分量为单位方差
            DATAx_normalized_y = pd.DataFrame(normalizer_y, columns=["list_GPP", "list_SIF"])
            # r,p = stats.pearsonr(list_x, list_y)
            pca1 = PCA()
            pca1.fit(DATAx_normalized_y)  # 训练模型
            vector1 = pca1.components_  # 返回模型的特征向量
            #vector_value = pca1.explained_variance_  # 返回模型的特征值
            b_y = pca1.explained_variance_ratio_  # 返回各成分的方差百分比
            C_y = np.dot(b_y,vector1)
            F_y = np.sum(C_y * np.array(DATAx_normalized_y), axis=1)

            R11, a = stats.pearsonr(F_y, F_x)

            std_normalized_x=F_x.std()
            std_normalized_y = F_y.std()
            C_x_diag=np.diag(C_x)
            C_y_diag = np.diag(C_y)
            R_1=np.dot(C_y_diag,r)
            R_2 = np.dot(R_1, C_x_diag)
            R = R_2 / (std_normalized_x*std_normalized_y)
            listR_GPP_PAR.append(R[0,0])
            listR_GPP_TMP.append(R[0,1])
            listR_GPP_PRE.append(R[0,2])
            listR_GPP_VPD.append(R[0,3])
            listR_GPP_SM.append(R[0,4])
            listR_SIF_PAR.append(R[1,0])
            listR_SIF_TMP.append(R[1,1])
            listR_SIF_PRE.append(R[1,2])
            listR_SIF_VPD.append(R[1,3])
            listR_SIF_SM.append(R[1,4])
            listR_GPP.append(np.sum(R,axis=1)[0])
            listR_SIF.append(np.sum(R,axis=1)[1])
            listR_PAR.append(np.sum(R,axis=0)[0])
            listR_TMP.append(np.sum(R,axis=0)[1])
            listR_PRE.append(np.sum(R,axis=0)[2])
            listR_VPD.append(np.sum(R,axis=0)[3])
            listR_SM.append(np.sum(R,axis=0)[4])
            listR.append(np.sum(R))
            listR11.append(R11)
        except:
            print("***********")
            rsquared = None
            listR_GPP_PAR.append(rsquared)
            listR_GPP_TMP.append(rsquared)
            listR_GPP_PRE.append(rsquared)
            listR_GPP_VPD.append(rsquared)
            listR_GPP_SM.append(rsquared)
            listR_SIF_PAR.append(rsquared)
            listR_SIF_TMP.append(rsquared)
            listR_SIF_PRE.append(rsquared)
            listR_SIF_VPD.append(rsquared)
            listR_SIF_SM.append(rsquared)
            listR_GPP.append(rsquared)
            listR_SIF.append(rsquared)
            listR_PAR.append(rsquared)
            listR_TMP.append(rsquared)
            listR_PRE.append(rsquared)
            listR_VPD.append(rsquared)
            listR_SM.append(rsquared)
            listR.append(rsquared)
            listR11.append(rsquared)
    else:
        rsquared =  None
        listR_GPP_PAR.append(rsquared)
        listR_GPP_TMP.append(rsquared)
        listR_GPP_PRE.append(rsquared)
        listR_GPP_VPD.append(rsquared)
        listR_GPP_SM.append(rsquared)
        listR_SIF_PAR.append(rsquared)
        listR_SIF_TMP.append(rsquared)
        listR_SIF_PRE.append(rsquared)
        listR_SIF_VPD.append(rsquared)
        listR_SIF_SM.append(rsquared)
        listR_GPP.append(rsquared)
        listR_SIF.append(rsquared)
        listR_PAR.append(rsquared)
        listR_TMP.append(rsquared)
        listR_PRE.append(rsquared)
        listR_VPD.append(rsquared)
        listR_SM.append(rsquared)
        listR.append(rsquared)
        listR11.append(rsquared)

book=openpyxl.Workbook()
sheet=book.create_sheet(index=0)
sheet.cell(column=1, row=1, value='FID')
sheet.cell(column=2, row=1, value='listR_GPP_PAR')
sheet.cell(column=3, row=1, value='listR_GPP_TMP')
sheet.cell(column=4, row=1, value='listR_GPP_PRE')
sheet.cell(column=5, row=1, value='listR_GPP_VPD')
sheet.cell(column=6, row=1, value='listR_GPP_SM')
sheet.cell(column=7, row=1, value='listR_SIF_PAR')
sheet.cell(column=8, row=1, value='listR_SIF_TMP')
sheet.cell(column=9, row=1, value='listR_SIF_PRE')
sheet.cell(column=10, row=1, value='listR_SIF_VPD')
sheet.cell(column=11, row=1, value='listR_SIF_SM')
sheet.cell(column=12, row=1, value='listR_GPP')
sheet.cell(column=13, row=1, value='listR_SIF')
sheet.cell(column=14, row=1, value='listR_PAR')
sheet.cell(column=15, row=1, value='listR_TMP')
sheet.cell(column=16, row=1, value='listR_PRE')
sheet.cell(column=17, row=1, value='listR_VPD')
sheet.cell(column=18, row=1, value='listR_SM')
sheet.cell(column=19, row=1, value='listR')
sheet.cell(column=20, row=1, value='listR11')

for m in range(len(listR_GPP)):
    print(m)
    sheet.cell(column=1, row=m + 2, value=FID[m])
    sheet.cell(column=2, row=m + 2, value=listR_GPP_PAR[m])
    sheet.cell(column=3, row=m + 2, value=listR_GPP_TMP[m])
    sheet.cell(column=4, row=m + 2, value=listR_GPP_PRE[m])
    sheet.cell(column=5, row=m + 2, value=listR_GPP_VPD[m])
    sheet.cell(column=6, row=m + 2, value=listR_GPP_SM[m])
    sheet.cell(column=7, row=m + 2, value=listR_SIF_PAR[m])
    sheet.cell(column=8, row=m + 2, value=listR_SIF_TMP[m])
    sheet.cell(column=9, row=m + 2, value=listR_SIF_PRE[m])
    sheet.cell(column=10, row=m + 2, value=listR_SIF_VPD[m])
    sheet.cell(column=11, row=m + 2, value=listR_SIF_SM[m])
    sheet.cell(column=12, row=m + 2, value=listR_GPP[m])
    sheet.cell(column=13, row=m + 2, value=listR_SIF[m])
    sheet.cell(column=14, row=m + 2, value=listR_PAR[m])
    sheet.cell(column=15, row=m + 2, value=listR_TMP[m])
    sheet.cell(column=16, row=m + 2, value=listR_PRE[m])
    sheet.cell(column=17, row=m + 2, value=listR_VPD[m])
    sheet.cell(column=18, row=m + 2, value=listR_SM[m])
    sheet.cell(column=19, row=m + 2, value=listR[m])
    sheet.cell(column=20, row=m + 2, value=listR11[m])


book.save(r'E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\CRU\comprehensive corr\VPM_GOME_R.xlsx')
print()
