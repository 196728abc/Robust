# -*- coding: UTF-8 -*-

import glob
import numpy as np
import pandas as pd
import os
import xlwt
from collections import Counter


import arcpy
from arcpy.sa import *
arcpy.CheckOutExtension('Spatial')

Points_shp = r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\flux_tower\data for soil moisture\GLOBAL DATA"
#GPP=glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\FLUXCOM", "*.tif"))
#SIF = glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\GOMESAT", "*.tif"))
#NDVI = glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\fesc\MODIS_fesc_NDVI_NIR\NDVI", "*.tif"))
#EVI = glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\fesc\MODIS_fesc_NDVI_NIR\EVI", "*.tif"))
#NIRv = glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\fesc\MODIS_fesc_NDVI_NIR\NIRv", "*.tif"))

#GPP=glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\VPM_GPP_Monthly_0_5Deg\GPP_0.5deg", "*.tif"))
SIF=glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\CSIF\111", "*.tif"))

SWR=glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\SWR", "*.tif"))
#TMP=glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\TMP", "*.tif"))
#PRE=glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\PRE", "*.tif"))
VPD=glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\VPD", "*.tif"))
SM=glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\TEST\SM", "*.tif"))
fPAR = glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\fesc\fesc_AVHRR\fPAR_AVHRR", "*.tif"))

#par = glob.glob(os.path.join(r"C:\\Users\lijia\Desktop\TEST\SWR", "*.tif"))
#tmp = glob.glob(os.path.join(r"C:\\Users\lijia\Desktop\TEST\TMP", "*.tif"))
#pre = glob.glob(os.path.join(r"C:\\Users\lijia\Desktop\TEST\PRE", "*.tif"))
#vpd = glob.glob(os.path.join(r"C:\\Users\lijia\Desktop\TEST\VPD", "*.tif"))
#sm = glob.glob(os.path.join(r"E:\\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\meoto\CRU\sm", "*.tif"))

#landcovers=glob.glob(os.path.join(r"E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\landCover\North hemsphere", "*.tif"))
#landcovers=[landcoverss[0],landcoverss[1],landcoverss[3],landcoverss[4],landcoverss[5],landcoverss[6],landcoverss[7],landcoverss[8]]
#NAME_lands=['CRO','DBF','EBF','ENF','GRA','MF','SAV','SHR']

ras=SIF[8]
kk=1
nameT = 'CSIF_meteo.shp'
outname = os.path.join(Points_shp, nameT)  # 合并输出文件名+输出路径
out_point1 = arcpy.RasterToPoint_conversion(ras, outname, "VALUE")

arcpy.CheckOutExtension("ImageAnalyst")  # 检查许可

for num_land in range(0,1):
    num_1=range(0,192)
    lists_1=[]
    str_list=[]
    for i in num_1:
        flux_gpp=SIF[i]
        #outExtractByMask = ExtractByMask(flux_gpp, landcovers[num_land])
        str = os.path.basename(flux_gpp)
        #str = str.rsplit('.', 1)[0]
        str ='SF_'+str[4:11]
        # str='M_'+str
        str_list.append(str)
        AA=[flux_gpp,str]
        lists_1.append(AA)

    if kk==1:
        num_2=range(0,192)
        lists_2=[]
        #str_list_2=[]
        for i in num_2:
            gosif=SWR[i]
            #outExtractByMask = ExtractByMask(gosif, landcovers[num_land])
            str = os.path.basename(gosif)
            str = 'SW_'+str[4:11]
           # str = str.rsplit('.', 1)[0]
            str_list.append(str)
            AA=[gosif,str]
            lists_2.append(AA)

        num_3 = range(0,192)
        lists_3 = []
        # str_list_2=[]
        for i in num_3:
            gosif = VPD[i]
            #outExtractByMask = ExtractByMask(gosif, landcovers[num_land])
            str = os.path.basename(gosif)
            # str_year = str.split(".")[0][18:22]  # GOSIF: str_year = str.split(".")[0][10:].  GOME2: str_year = str.split(".")[0][18:22]. TROPO: str_year = str.split("_")[2]
            # str_month = str.split(".")[0][22:24]     # GOSIF:  str_month = str.split(".")[1].  GOME2: str_month = str.split(".")[0][22:24]. TROPO: str_month = str.split("_")[3][0:2]
            # str_1 = str_year + "_" + str_month
            # str = str.rsplit('.', 1)[0]
            str = "VD"+str[4:11]
            str_list.append(str)
            AA = [gosif, str]
            lists_3.append(AA)

        num_4 = range(0,192)
        lists_4 = []
        # str_list_2=[]
        for i in num_4:
            gosif = SM[i]
            #  outExtractByMask = ExtractByMask(gosif, landcovers[num_land])
            str = os.path.basename(gosif)
            # str_year = str.split(".")[0][18:22]  # GOSIF: str_year = str.split(".")[0][10:].  GOME2: str_year = str.split(".")[0][18:22]. TROPO: str_year = str.split("_")[2]
            # str_month = str.split(".")[0][22:24]     # GOSIF:  str_month = str.split(".")[1].  GOME2: str_month = str.split(".")[0][22:24]. TROPO: str_month = str.split("_")[3][0:2]
            # str_1 = str_year + "_" + str_month
            # str = str.rsplit('.', 1)[0]
            str = "SM_"+str[3:10]
            str_list.append(str)
            AA = [gosif, str]
            lists_4.append(AA)

        num_5 = range(0, 192)
        lists_5 = []
        # str_list_2=[]
        for i in num_5:
            gosif = fPAR[i]
            #  outExtractByMask = ExtractByMask(gosif, landcovers[num_land])
            str = os.path.basename(gosif)
            # str_year = str.split(".")[0][18:22]  # GOSIF: str_year = str.split(".")[0][10:].  GOME2: str_year = str.split(".")[0][18:22]. TROPO: str_year = str.split("_")[2]
            # str_month = str.split(".")[0][22:24]     # GOSIF:  str_month = str.split(".")[1].  GOME2: str_month = str.split(".")[0][22:24]. TROPO: str_month = str.split("_")[3][0:2]
            # str_1 = str_year + "_" + str_month
            # str = str.rsplit('.', 1)[0]
            str = "fR_" + str[5:12]
            str_list.append(str)
            AA = [gosif, str]
            lists_5.append(AA)
    else:
        num_2 = range(0,192)
        lists_2 = []
        # str_list_2=[]
        for i in num_2:
            gosif = VPD[i]
            #  outExtractByMask = ExtractByMask(gosif, landcovers[num_land])
            str = os.path.basename(gosif)
            # str_year = str.split(".")[0][18:22]  # GOSIF: str_year = str.split(".")[0][10:].  GOME2: str_year = str.split(".")[0][18:22]. TROPO: str_year = str.split("_")[2]
            # str_month = str.split(".")[0][22:24]     # GOSIF:  str_month = str.split(".")[1].  GOME2: str_month = str.split(".")[0][22:24]. TROPO: str_month = str.split("_")[3][0:2]
            # str_1 = str_year + "_" + str_month
            # str = str.rsplit('.', 1)[0]
            str = "VD_" + str[4:11]
            str_list.append(str)
            AA = [gosif, str]
            lists_2.append(AA)

        num_3 = range(0,192)
        lists_3 = []
        # str_list_2=[]
        for i in num_3:
            gosif = SM[i]
            #  outExtractByMask = ExtractByMask(gosif, landcovers[num_land])
            str = os.path.basename(gosif)
            # str_year = str.split(".")[0][18:22]  # GOSIF: str_year = str.split(".")[0][10:].  GOME2: str_year = str.split(".")[0][18:22]. TROPO: str_year = str.split("_")[2]
            # str_month = str.split(".")[0][22:24]     # GOSIF:  str_month = str.split(".")[1].  GOME2: str_month = str.split(".")[0][22:24]. TROPO: str_month = str.split("_")[3][0:2]
            # str_1 = str_year + "_" + str_month
            # str = str.rsplit('.', 1)[0]
            str = "SM_" + str[3:10]
            str_list.append(str)
            AA = [gosif, str]
            lists_3.append(AA)

    ExtractMultiValuesToPoints(out_point1, lists_1, "NONE")
    ExtractMultiValuesToPoints(out_point1, lists_2, "NONE")
    ExtractMultiValuesToPoints(out_point1, lists_3, "NONE")
    ExtractMultiValuesToPoints(out_point1, lists_4, "NONE")
    ExtractMultiValuesToPoints(out_point1, lists_5, "NONE")

    #ExtractMultiValuesToPoints(out_point1, lists_5, "NONE")
    #ExtractMultiValuesToPoints(out_point1, lists_6, "NONE")
    #ExtractMultiValuesToPoints(out_point1, lists_7, "NONE")

print()




'''
from sklearn.linear_model import LinearRegression
import openpyxl as op
import scipy.stats as stats
from matplotlib.font_manager import FontProperties
from statsmodels.formula.api import ols
#from sklearn.preprocessing import StandardScaler
import statsmodels.api as sm
import openpyxl

def standardize(x):
   # return (x - np.mean(x))/(np.std(x))
    return (x - np.min(x))/(np.max(x)-np.min(x))

data = pd.read_csv(r"C:\Users\lijia\Desktop\TEST\Points\VPM_TROPOextend_1.csv",header=0,dtype=np.float64)
data=np.array(data)

kk=192
data_1=np.where(data[:,3:3+kk]>-5,data[:,3:3+kk],None)
data_2=np.where(data[:,3+kk:3+2*kk]>-5,data[:,3+kk:3+2*kk],None)
#data_2=np.where(data[:,243:363]>0,data[:,243:363],None)
#data_2=np.where(data[:,363:]>0,data[:,363:],None)


data_T=np.append(data_1,data_2,axis=1)
#data_T=np.append(data_T,data_3,axis=1)
#data_T=np.append(data_T,data_4,axis=1)
FID=data[:,0]

listR = []
listC = []
listP = []
listK = []


for i in range (0,data_T.shape[0]):
    listy = data_T[i][0:kk]
    listx = data_T[i][kk:]
   # listz = data_T[i][240:360]
   # listf = data_T[i][360:]
    list_y=[]
    list_x=[]
   # list_z = []
   # list_f = []
    for j in range(0,kk):
        if listy[j]!=None and listx[j]!=None:# and listz[j]!=None and listf[j]!=None:
            list_y.append(listy[j])
            list_x.append(listx[j])
       #     list_z.append(listz[j])
        #    list_f.append(listf[j])
    print(i)
  #  list_y=np.array(list_y)/np.array(list_x)
  #  list_x = np.array(list_z)/np.array(list_f)

    if len(list_y)>4:
        #list_x = standardize(list_x)
        #list_y = standardize(list_y)
        list_x=list_x
        list_y = list_y
        list_x = sm.add_constant(list_x)
        mod = sm.OLS(list_y, list_x)
        result = mod.fit()
        # model = ols('y~x', data1).fit()
        if result.f_pvalue > 0.05 or np.isnan(result.f_pvalue):
            rsquared = params = pvalues = interpct = None
            listR.append(0)
            listC.append(params)
            listP.append(pvalues)
            listK.append(interpct)

        else:
            rsquared = result.rsquared
            interpct=result.params[0]
            params = result.params[1]
            pvalues = result.f_pvalue
            listR.append(rsquared)
            listC.append(params)
            listP.append(pvalues)
            listK.append(interpct)

    else:
        rsquared = params = pvalues = interpct = None
        listR.append(rsquared)
        listC.append(params)
        listP.append(pvalues)
        listK.append(interpct)

book=openpyxl.Workbook()
sheet=book.create_sheet(index=0)
sheet.cell(column=1, row=1, value='FID')
sheet.cell(column=2, row=1, value='List_R')
sheet.cell(column=3, row=1, value='List_C')
sheet.cell(column=4, row=1, value='List_P')
sheet.cell(column=5, row=1, value='List_K')

for m in range(len(listR)):
    print(m)
    sheet.cell(column=1, row=m + 2, value=FID[m])
    sheet.cell(column=2, row=m + 2, value=listR[m])
    sheet.cell(column=3, row=m + 2, value=listC[m])
    sheet.cell(column=4, row=m + 2, value=listP[m])
    sheet.cell(column=5, row=m + 2, value=listK[m])

book.save(r'E:\1_canopy_photosynthesis\photosynthesis_and_Global_Climate\SIF\data\GPP_SIF\VPM_TROPOextend_linear_reg.xlsx')
print()
'''